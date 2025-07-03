import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Step 1: Convert CSV to Excel (XLSX)
csv_file = "Mar ledger for Jun25 report.csv"  # Replace with your CSV file name
excel_file = "08Jun25Report.xlsx"  # Ensure this ends with .xlsx

# Read the CSV into a DataFrame
df = pd.read_csv(csv_file)

# Clean `[Ledger] Staff` column: Fill missing values, trim spaces
df["[Ledger] Staff"] = df["[Ledger] Staff"].fillna("").str.strip()

# Convert [Ledger] Time to hours
if "[Ledger] Time" in df.columns:
    def convert_time_to_hours(time_str):
        try:
            hours, minutes = map(int, time_str.split(":"))
            return hours + minutes / 60
        except:
            return 0  # Default to 0 for invalid entries

    df["Ledger Time Hours"] = df["[Ledger] Time"].apply(convert_time_to_hours)

# Convert `[Ledger] Invoiced Amount` and `[Ledger] Billable Amount` to numeric
df["[Ledger] Invoiced Amount"] = pd.to_numeric(df["[Ledger] Invoiced Amount"], errors="coerce").fillna(0)
df["[Ledger] Billable Amount"] = pd.to_numeric(df["[Ledger] Billable Amount"], errors="coerce").fillna(0)

# Save the cleaned data to Excel
with pd.ExcelWriter(excel_file, engine="openpyxl", mode="w") as writer:
    df.to_excel(writer, index=False, sheet_name="Data")

print("Step 1: Cleaned data saved to Excel.")

# Step 2: Define Billable Details for All Staff
staff_billable_details = pd.DataFrame({
    "Staff": ["Jasmine Kee",	"Johncel Lai",	"Kelly Wong",	"Madelyn Morales",	"Marianne Tejedor",	"Rachel Teoh",	"Stephen  Tejedor"	,"Sydelle Tay",	"Ted Soo"
],
    "Billable Percentage": [0.8,0.65,0.75,  0.75,   0.2,    0.8 ,0.2,   0.4,    0.70],
    "Billable Rate": [80    ,170,   135 ,80 ,60,    80  ,40,    220,    135]
})

# Excluded Staff
excluded_staff = ["Mei Tu", "Teke Tu"]

# Step 3: Loop through all unique staff names and calculate
unique_staff = df["[Ledger] Staff"].unique()  # Get all unique staff names
detailed_summary = []  # Store detailed summary calculations for all staff
summary_table = []  # Store summary table

for staff in unique_staff:
    if staff and staff not in excluded_staff:  # Skip empty or excluded staff
        # 1. Sum of `[Ledger] Invoiced Amount` for this staff
        sum_invoiced_amount = df[df["[Ledger] Staff"] == staff]["[Ledger] Invoiced Amount"].sum()

        # 2. Sum of Ledger Time Hours for this staff
        sum_ledger_time_hours = df[df["[Ledger] Staff"] == staff]["Ledger Time Hours"].sum()

        # 3. Sum of Ledger Time Hours for this staff AND `[Job] Job No.` = "J003167"
        sum_ledger_time_job_staff = df[
            (df["[Ledger] Staff"] == staff) & (df["[Job] Job No."] == "J003167")
        ]["Ledger Time Hours"].sum()

        # 4. Total Working Hours = Total Time Hours - Time for Job J003167
        total_working_hours = sum_ledger_time_hours - sum_ledger_time_job_staff

        # 5. Manual Calculation: Target Billable Amount
        staff_details = staff_billable_details[staff_billable_details["Staff"] == staff]
        if not staff_details.empty:
            billable_percentage = staff_details["Billable Percentage"].iloc[0]
            billable_rate = staff_details["Billable Rate"].iloc[0]
            target_billable_amount = total_working_hours * billable_percentage * billable_rate
        else:
            billable_percentage = 0
            billable_rate = 0
            target_billable_amount = 0

        # 6. Sum of `[Ledger] Billable Amount` for this staff
        sum_billable_amount = df[df["[Ledger] Staff"] == staff]["[Ledger] Billable Amount"].sum()

        # 7. Add "Target Hit?" logic
        target_hit = "Yes" if target_billable_amount < sum_invoiced_amount else "No"

        # 8. Add "Billable Hit?" logic
        billable_hit = "Yes" if target_billable_amount < sum_billable_amount else "No"

        # Append detailed rows for the staff
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Sum of [Ledger] Invoiced Amount for {staff}",
            "Value": sum_invoiced_amount
        })
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Sum of Ledger Time Hours for {staff}",
            "Value": sum_ledger_time_hours
        })
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Sum of Ledger Time Hours for {staff} AND Job No. J003167",
            "Value": sum_ledger_time_job_staff
        })
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Total Working Hours for {staff}",
            "Value": total_working_hours
        })
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Target Billable Amount for {staff}",
            "Value": target_billable_amount
        })
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Sum of [Ledger] Billable Amount for {staff}",
            "Value": sum_billable_amount
        })
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Target Hit? {staff}",
            "Value": target_hit
        })
        detailed_summary.append({
            "Staff": staff,
            "Calculation": f"Billable Hit? {staff}",
            "Value": billable_hit
        })

        # Append to summary table
        summary_table.append({
            "Staff": staff,
            "Target Hit": target_hit,
            "Billable Hit": billable_hit
        })
# Step 4: Save Detailed Summary and Summary Table to Excel
with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a") as writer:
    # Save detailed summary
    detailed_summary_df = pd.DataFrame(detailed_summary)
    detailed_summary_df.to_excel(writer, index=False, sheet_name="Detailed Summary")

    # Save summary table
    summary_table_df = pd.DataFrame(summary_table)
    summary_table_df.to_excel(writer, index=False, sheet_name="Summary Table")

# Step 5: Apply Conditional Formatting to Summary Table
wb = load_workbook(excel_file)
ws_summary = wb["Summary Table"]

# Define fills for conditional formatting
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Loop through rows in the Summary Table sheet
for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, min_col=2, max_col=3):  # Columns B and C
    for cell in row:
        if cell.value == "Yes":
            cell.fill = green_fill
        elif cell.value == "No":
            cell.fill = red_fill

# Save the workbook
wb.save(excel_file)
print(f"Step 5: Conditional formatting applied and saved to '{excel_file}'.")

import os
from openpyxl.styles import Font, Alignment

# Step 6: Create individual staff Excel files
excel_file = excel_file
summary_table = pd.read_excel(excel_file, sheet_name="Summary Table")
detailed_summary = pd.read_excel(excel_file, sheet_name="Detailed Summary")
data_tab = pd.read_excel(excel_file, sheet_name="Data")

for staff in unique_staff:
    if staff and staff not in excluded_staff:  # Skip empty or excluded staff
        # Filter the ledger data for this staff
        staff_ledger = df[df["[Ledger] Staff"] == staff]

        # Prepare a detailed summary for this staff
        filtered_data = data_tab[data_tab["[Ledger] Staff"] == staff]

    # 2. Filter the Detailed Summary for this staff
        filtered_summary = detailed_summary[detailed_summary["Staff"] == staff]
        
        # Create a new Excel file
        staff_excel_file = f"{staff}_performance_report.xlsx"
        with pd.ExcelWriter(staff_excel_file, engine="openpyxl") as writer:
            # Add the Data Page
            filtered_data.to_excel(writer, index=False, sheet_name="Data Page")

            # Add the Detailed Summary Page
            filtered_summary.to_excel(writer, index=False, sheet_name="Detailed Summary")

        # Apply Formatting to the Excel File
        wb = writer.book
        ws_summary = wb["Detailed Summary"]

        # Format the Detailed Summary Page
        for cell in ws_summary["1:1"]:  # Header row
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 50
        ws_summary.column_dimensions["C"].width = 15

        print(f"Customised Excel file created for {staff}: {staff_excel_file}")

    print("All files have been created.")
        

import pandas as pd
import yagmail

# Step 1: Load the Summary Table and Detailed Summary
summary_file = excel_file  # Your main summary Excel file
summary_table = pd.read_excel(summary_file, sheet_name="Summary Table")
detailed_summary = pd.read_excel(summary_file, sheet_name="Detailed Summary")

# Step 2: Load the Email List
email_file = "staff_emails.csv"  # CSV file with columns: Staff, Email
email_list = pd.read_csv(email_file)

# Step 3: Merge the Email List with the Summary Table
summary_table = summary_table.merge(email_list, on="Staff", how="left")

# Email Configuration
sender_email = "@gmail.com"  # Replace with your Gmail address
app_password = ""  # Replace with your Gmail App Password
yag = yagmail.SMTP(user=sender_email, password=app_password)

# Step 4: Iterate Over Staff and Send Emails
for index, row in summary_table.iterrows():
    staff = row["Staff"]
    email = row["Email"]  # This is now from the merged table
    target_hit = row["Target Hit"]

    # Skip if Email is Missing
    if pd.isna(email):
        print(f"No email found for {staff}. Skipping...")
        continue

    # Step 5: Filter Data for the Staff
    staff_ledger = df[df["[Ledger] Staff"] == staff]
    filtered_data = detailed_summary[detailed_summary["Staff"] == staff]

    # Step 6: Generate a Customised Report
    staff_file = f"{staff}_performance_report.xlsx"
    with pd.ExcelWriter(staff_file, engine="openpyxl") as writer:
        
        filtered_data.to_excel(writer, index=False, sheet_name="Data Page")

            # Add the Detailed Summary Page
        staff_ledger.to_excel(writer, index=False, sheet_name="Detailed Summary")
        
    # Step 7: Prepare Email Content
    if target_hit == "Yes":
        subject = f"Congratulations on Hitting Your Target, {staff}!"
        body = f"""
        Hi {staff},

        Congratulations on hitting your performance target! 🎉
        Your hard work and dedication have paid off, and we’re thrilled with your success.

        Attached is your performance report for this period. Keep up the great work!

        Regards,
        Performance Team
        """
    else:
        subject = f"Performance Feedback for {staff}"
        body = f"""
        Hi {staff},

        We noticed you didn’t hit your performance target this time. We believe in your potential, and we're here to support you in achieving your goals.

        Attached is your performance report for this period. Please review the report, and let us know if there’s any way we can assist you.

        Regards,
        Performance Team
        """

    # Step 8: Send Email
    try:
        yag.send(to=email, subject=subject, contents=body, attachments=staff_file)
        print(f"Email sent to {staff}: {email}")
    except Exception as e:
        print(f"Failed to send email to {staff}: {str(e)}")

print("All emails have been sent.")
